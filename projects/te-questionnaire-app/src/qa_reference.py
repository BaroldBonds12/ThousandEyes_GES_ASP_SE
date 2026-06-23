"""
QA Reference — 3-column Excel spreadsheet saved to the local cache directory.

When the user saves an answered file, this module writes every Q&A pair to
~/.te_qa_cache/qa_reference.xlsx.  The processor loads it on startup and
injects matching rows as additional context so Ollama can reference past
human-reviewed answers.
"""

from __future__ import annotations

import difflib
from pathlib import Path
from typing import Any

REFERENCE_PATH = Path.home() / ".te_qa_cache" / "qa_reference.xlsx"

# Fuzzy-match threshold for context injection
_MATCH_THRESHOLD = 0.68


# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------

def save_reference(
    questions: list[dict[str, Any]],
    path: Path = REFERENCE_PATH,
) -> int:
    """
    Write all answered Q&A pairs to a styled 3-column Excel spreadsheet.

    Columns: Question | Answer | Source URL

    Returns the number of Q&A rows written.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QA Reference"

    # ── Header row ────────────────────────────────────────────────────────
    hdr_fill = PatternFill(start_color="2E2E50", end_color="2E2E50", fill_type="solid")
    hdr_font = Font(color="A78BFF", bold=True, size=11)
    hdr_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for col_idx, header in enumerate(["Question", "Answer", "Source URL"], 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = hdr_align

    # Freeze the header row so it stays visible when scrolling
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────
    even_fill = PatternFill(start_color="1E1E2E", end_color="1E1E2E", fill_type="solid")
    odd_fill  = PatternFill(start_color="22223A", end_color="22223A", fill_type="solid")
    data_font  = Font(color="E0E0F0", size=10)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_side  = Side(border_style="thin", color="2E2E50")
    thin_border = Border(bottom=thin_side)

    answered = [q for q in questions if q.get("answer")]
    for i, q in enumerate(answered):
        row_num = i + 2
        question_text = (q.get("question") or q.get("rfp_requirement") or "").strip()
        answer_text   = (q.get("answer") or "").strip()
        url_text      = (q.get("source_url") or "").strip()

        fill = even_fill if i % 2 == 0 else odd_fill
        for col_idx, val in enumerate([question_text, answer_text, url_text], 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill      = fill
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = thin_border

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 52

    # ── Row heights ───────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 20
    for i in range(len(answered)):
        ws.row_dimensions[i + 2].height = 45  # tall enough for 3-4 lines

    wb.save(path)
    return len(answered)


# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------

def load_reference(path: Path = REFERENCE_PATH) -> list[dict[str, str]]:
    """
    Load the reference spreadsheet.

    Returns a list of dicts:  {"question": str, "answer": str, "source_url": str}
    Returns an empty list if the file doesn't exist or can't be read.
    """
    if not path.exists():
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            q, a, u = (list(row) + [None, None, None])[:3]
            if q and a:
                rows.append({
                    "question":   str(q).strip(),
                    "answer":     str(a).strip(),
                    "source_url": str(u).strip() if u else "",
                })
        wb.close()
        return rows
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Context injection helper (used by processor.py)
# ---------------------------------------------------------------------------

def find_reference_context(question: str, ref_rows: list[dict[str, str]]) -> str:
    """
    Return a short context block of the most similar reference Q&A entries
    to inject into the LLM prompt.  Returns empty string if no good match.
    """
    if not ref_rows:
        return ""

    norm_q = " ".join(question.lower().split())
    scored: list[tuple[float, dict]] = []

    for row in ref_rows:
        ratio = difflib.SequenceMatcher(
            None, norm_q, " ".join(row["question"].lower().split())
        ).ratio()
        if ratio >= _MATCH_THRESHOLD:
            scored.append((ratio, row))

    if not scored:
        return ""

    scored.sort(key=lambda x: x[0], reverse=True)
    top = scored[:3]

    lines = ["[Relevant answers from previous human-reviewed sessions:]"]
    for _, row in top:
        lines.append(f"Q: {row['question']}")
        lines.append(f"A: {row['answer']}")
        if row.get("source_url"):
            lines.append(f"Source: {row['source_url']}")
        lines.append("")

    return "\n".join(lines).strip()
